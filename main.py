import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from openpyxl import load_workbook
import os
import win32print
import win32api
import traceback


class ExcelFillerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 模板填充工具")
        self.root.geometry("500x300")

        self.selected_template = None  # 当前选择的模板

        # 固定数据来源
        self.data_source = {
            "name": "张三",
            "idNumber": "123456789012345678",
            "address": "北京市朝阳区幸福路1号",
            "phone": "18812345678",
            "birthDate": "1990-01-01"
        }

        # 界面布局
        self.upload_template_button = tk.Button(root, text="上传模板文件", command=self.upload_template)
        self.upload_template_button.pack(pady=10)

        self.template_label = tk.Label(root, text="已上传模板: 无", fg="blue")
        self.template_label.pack(pady=5)

        self.fill_and_print_button = tk.Button(root, text="填充并打印", command=self.fill_and_print)
        self.fill_and_print_button.pack(pady=10)

    def upload_template(self):
        """上传模板文件"""
        try:
            file = filedialog.askopenfile(
                title="选择Excel模板文件",
                filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xlsb")],
                mode="r"  # 打开文件以只读模式
            )
            if file:
                self.selected_template = file.name  # 获取文件路径
                self.template_label.config(text=f"已上传模板: {os.path.basename(self.selected_template)}")
                file.close()
        except Exception as e:
            messagebox.showerror("错误", f"上传模板文件失败: {e}")

    def fill_and_print(self):
        """填充模板并打印"""
        if not self.selected_template:
            messagebox.showerror("错误", "请选择模板文件！")
            return

        # 加载模板
        try:
            wb = load_workbook(self.selected_template)
            sheet = wb.active

            # 遍历所有单元格，查找占位符并替换
            for r in sheet.iter_rows():
                for cell in r:
                    if isinstance(cell.value, str) and "{{" in cell.value and "}}" in cell.value:
                        placeholder = cell.value.strip("{{").strip("}}")  # 去掉 {{ 和 }}
                        if placeholder in self.data_source:
                            cell.value = self.data_source[placeholder]  # 替换为固定数据中的值

            # 保存临时文件
            temp_file = os.path.join(os.getcwd(), "filled_template.xlsx")
            wb.save(temp_file)

            # 调用打印机（Windows 上使用 win32print 和 win32api 打印）
            self.print_file(temp_file)

            messagebox.showinfo("成功", "模板填充完成，已发送到打印机！")
        except Exception as e:
            messagebox.showerror("错误", f"填充或打印失败: {e}")

    def print_file(self, filepath):
        """使用默认打印机打印文件"""
        try:
            # 获取默认打印机
            printer_name = win32print.GetDefaultPrinter()
            print(f"使用打印机: {printer_name}")

            # 使用默认打印机发送文件
            win32api.ShellExecute(0, "print", filepath, None, ".", 0)
        except Exception as e:
            messagebox.showerror("错误", f"打印失败: {e}")

# 捕获未处理异常
def exception_handler(exc_type, exc_value, exc_tb):
    error_message = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    messagebox.showerror("错误", f"程序发生异常:\n{error_message}")

# 设置全局异常处理
import sys
sys.excepthook = exception_handler

# 创建应用程序
root = tk.Tk()
app = ExcelFillerApp(root)
root.mainloop()